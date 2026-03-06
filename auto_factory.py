import streamlit as st
import pandas as pd
import os
import openpyxl
from PIL import Image
import re
from datetime import datetime

# 방금 분리한 3개의 모듈을 통째로 불러옵니다!
from module_api import *
from module_data import *
from module_image import *

st.set_page_config(layout="wide", page_title="초정밀 소싱 판독기 v10.9 (무결점 3구역 시스템)")

# --- [상태 관리] ---
if 'parsed_data' not in st.session_state: st.session_state.parsed_data = None
if 'api_error_log' not in st.session_state: st.session_state.api_error_log = None
if 'calculated_skus' not in st.session_state: st.session_state.calculated_skus = []
if 'preview_imgs' not in st.session_state: st.session_state.preview_imgs = {} 
if 'ai_size_json' not in st.session_state: st.session_state.ai_size_json = None 
if 'needs_size_chart' not in st.session_state: st.session_state.needs_size_chart = False 

if 'folder_list' not in st.session_state: st.session_state.folder_list = []
if 'current_folder_idx' not in st.session_state: st.session_state.current_folder_idx = 0
if 'batch_logs' not in st.session_state: st.session_state.batch_logs = []
if 'url_list' not in st.session_state: st.session_state.url_list = []
if 'current_url_idx' not in st.session_state: st.session_state.current_url_idx = 0

def reset_state(): 
    st.session_state.parsed_data = None
    st.session_state.api_error_log = None
    st.session_state.calculated_skus = []
    st.session_state.preview_imgs = {}
    st.session_state.ai_size_json = None
    st.session_state.folder_list = []
    st.session_state.current_folder_idx = 0
    st.session_state.url_list = []
    st.session_state.current_url_idx = 0
    st.session_state.batch_logs = []

# ==========================================
# 사이드바 & 3구역 작업실 자동 생성
# ==========================================
MY_API_KEY = "AIzaSyA_9S9Vw0Wz3M_pRhN6lgByXWgjZpvNU9s" 
MY_RAPID_KEY = "6f9986ef72msh6c1d9dc5468276ep1583f0jsn98c290f0645e"
today_str = datetime.now().strftime('%m%d')

# --- [스마트 경로 탐색기] 데스크탑/노트북 바탕화면 자동 인식 ---
user_home = os.path.expanduser("~")
desktop_path = os.path.join(user_home, "Desktop")
onedrive_desktop = os.path.join(user_home, "OneDrive", "바탕 화면")

# 만약 데스크탑처럼 OneDrive 바탕화면을 쓴다면 그곳으로 자동 연결
if os.path.exists(onedrive_desktop):
    desktop_path = onedrive_desktop

default_base = os.path.join(desktop_path, f"오늘의 소싱 {today_str}")
# -------------------------------------------------------------

default_local = os.path.join(default_base, "1. 로컬 폴더 모드 작업용")
default_api = os.path.join(default_base, "2. API URL 리스트 모드 작업용")
default_out = os.path.join(default_base, "3. 최종 결과물 제출용")

os.makedirs(default_local, exist_ok=True)
os.makedirs(default_api, exist_ok=True)
os.makedirs(default_out, exist_ok=True)

with st.sidebar:
    st.subheader("⚙️ 1. 기본 설정")
    api_key = st.text_input("🔑 Gemini API Key", type="password", value=MY_API_KEY)
    rapid_api_key = st.text_input("🔌 Rapid API Key (1688용)", type="password", value=MY_RAPID_KEY, help="URL 모드 사용 시 필수")
    brand_name = st.text_input("🏷️ 브랜드명", value="미화부장")
    
    st.markdown("---")
    st.subheader("🤖 2. 무인 자동화 공통 세팅")
    global_ex_rate = st.number_input("환율 (원/¥)", value=270, step=5)
    global_target_profit = st.number_input("목표 '최소' 순이익 (원)", value=8000, step=500)
    
    global_main_strategy = st.radio("자동화 대표 이미지 방식", ["[옵션 1] 원본 보존 1:1 패딩", "[옵션 2] AI 배경제거 & 스튜디오"], index=0)
    global_detail_strategy = st.radio("자동화 상세 이미지 방식", ["[옵션 1] 해당 썸네일 + 메인 병합", "[옵션 2] 설명 이미지 쫙 이어붙이기"], index=1)
    
    st.markdown("---")
    if st.button("🔄 전체 초기화", on_click=reset_state, use_container_width=True): pass

# ==========================================
# 🚀 무인 자동화 헤드리스 파이프라인
# ==========================================
def process_folder_fully_headless(folder_path, root_path, ex_rate, target_profit, main_strategy, detail_strategy):
    info_file, _ = find_target_txt(folder_path) 
    if not info_file: raise Exception("텍스트 파일(_URL.txt)이 없습니다.")
    excel_file = find_excel_template(root_path) or find_excel_template(folder_path)
    if not excel_file: raise Exception("쿠팡 카테고리 견적서(.xlsx)가 없습니다.")
    
    try:
        with open(info_file, 'r', encoding='utf-8') as f: raw_text = f.read()
    except:
        with open(info_file, 'r', encoding='cp949') as f: raw_text = f.read()
        
    d = parse_1688_text(raw_text, MY_API_KEY)
    ai, error_log = analyze_with_ai(MY_API_KEY, d['product_title'], raw_text)
    if error_log: raise Exception(f"AI 분석 실패: {error_log}")
    if ai: ai['tags'] = ai.get('tags', '').replace('#', '').strip()
    else: ai = {"title": "", "tags": "", "material": d['attributes'].get('소재', ''), "alt_text": "", "season": ""}
    
    df = pd.DataFrame(d['sku_list'])
    if df.empty: raise Exception("SKU(옵션) 데이터를 찾을 수 없습니다.")
    df['원가(원)'] = (df['위안화'] * ex_rate).astype(int)
    df[['최종 납품가(원)', '내 마진(원)', '쿠팡판매가(원)', '쿠팡마진율(%)', '권장소비자가(원)']] = df.apply(lambda row: optimize_profit_margin(row, target_profit), axis=1)
    calculated_skus = df.to_dict('records')
    
    for sku in calculated_skus:
        sku['옵션명'] = clean_option_name(sku['옵션명'], MY_API_KEY, sku.get('이미지', ''))
    
    suggested_name = ai.get('title', '')
    final_material, final_tags, alt_text, season = ai.get('material', ''), ai.get('tags', ''), ai.get('alt_text', ''), ai.get('season', '')
    pack_weight, pack_size = d['attributes'].get('무게', '500'), "300*400*50"
    
    # [수술 1] "고객센터: 고객센터" 중복 제거를 위해 순수 전화번호만 입력
    importer, phone = "경기도 시흥시 비둘기공원7길 65 유한프라자 502호 삼삼오오", "070-4128-6398"
    
    output_path = os.path.join(default_out, f"[{brand_name}] {get_clean_filename(brand_name, suggested_name, '')}_최종결과물")
    os.makedirs(output_path, exist_ok=True)
    
    columns, reqs, target_sheet_name = extract_excel_columns(excel_file)
    wb = openpyxl.load_workbook(excel_file)
    ws = wb[target_sheet_name]
    
    col_info_list = []
    for cell in ws[5]:
        if cell.value:
            c_name = str(cell.value).strip()
            c_req = str(ws.cell(row=6, column=cell.column).value).strip()
            col_info_list.append((cell.column, c_name, c_req))
            
    start_row = 9 
    needs_size_chart = any("사이즈표" in str(c) or "사이즈차트" in str(c) for c in columns)
    st.session_state.needs_size_chart = needs_size_chart
    
    for i, sku in enumerate(calculated_skus):
        current_row = start_row + i
        base_fname = get_clean_filename(brand_name, suggested_name, sku['옵션명'])
        
        for col_idx, col_name, col_req in col_info_list:
            val = get_excel_mapped_value(col_name, col_req, d, sku, base_fname, brand_name, suggested_name, final_tags, alt_text, season, pack_weight, pack_size)
            if val is not None and val != "nan" and val != "":
                ws.cell(row=current_row, column=col_idx, value=val)
                
    wb.save(os.path.join(output_path, f"{os.path.splitext(os.path.basename(excel_file))[0]}_제출.xlsx"))
    
    base_size = create_smart_size_chart(calculated_skus) if needs_size_chart else None
    for idx, sku in enumerate(calculated_skus):
        opt_name, base_fname = sku['옵션명'], get_clean_filename(brand_name, suggested_name, sku['옵션명'])
        c_color, c_size = split_color_size(opt_name)
        if sku['이미지']:
            di = download_img(sku['이미지'])
            if di: 
                p_main = create_studio_main_image(di) if "옵션 2" in main_strategy else trim_and_crop_1_to_1(di)
                p_main.save(os.path.join(output_path, f"{base_fname}.jpg"), format="JPEG", quality=100)
        
        display_opt = opt_name.replace("혼합색상", "").replace("단일색상", "").replace("혼합", "").replace(",", "").strip()
        display_opt = re.sub(r'\s+', ' ', display_opt)
        
        full_item_name = f"{brand_name} {suggested_name}"
        urls = get_final_detail_urls(folder_path, d, sku, detail_strategy)
        
        p_detail = create_dynamic_detail_page(full_item_name, display_opt, final_material, c_size, "중국", urls, brand_name)
        if p_detail: p_detail.save(os.path.join(output_path, f"{base_fname}detail.jpg"), format="JPEG", quality=85)
        p_label = create_perfect_korean_label_900x1200(f"{full_item_name} {display_opt}".strip(), c_size, final_material, importer, phone)
        p_label.save(os.path.join(output_path, f"{base_fname}label.jpg"), format="JPEG", quality=100)
        if needs_size_chart and base_size: base_size.save(os.path.join(output_path, f"{base_fname}size.jpg"), format="JPEG", quality=95)
            
    return {"folder_path": folder_path, "folder_name": os.path.basename(folder_path), "excel_path": excel_file, "ai_result": ai, "data": d, "raw_text": raw_text, "calculated_skus": calculated_skus, "output_path": output_path, "original_title": suggested_name}

def process_url_fully_headless(url, folder_path, template_name, rapid_key, gemini_key, ex_rate, target_profit, main_strategy, detail_strategy, custom_category=""):
    target_excel_path = None
    if template_name:
        search_keyword = template_name.replace('.xlsx', '').strip()
        try:
            for f in os.listdir(folder_path):
                if search_keyword in f and f.endswith('.xlsx') and not f.startswith('~$'):
                    target_excel_path = os.path.join(folder_path, f)
                    break
        except: pass
            
    if not target_excel_path: target_excel_path = find_excel_template(folder_path)
    if not target_excel_path: raise Exception(f"엑셀 템플릿(견적서)을 찾을 수 없습니다. (검색어: {template_name})")

    item_id = extract_item_id(url)
    if not item_id: raise Exception("유효한 1688 상품 ID를 찾을 수 없습니다.")
    
    api_item, api_err = fetch_1688_item_api(item_id, rapid_key)
    if not api_item: raise Exception(f"1688 API 에러: {api_err}")
    
    d = convert_api_to_v9_dict(api_item)
    if custom_category: d['category'] = custom_category 
        
    fake_raw_text = f"카테고리: {d['category']}\n제품 제목:\n{d['product_title']}\n"
    fake_raw_text += "메인 이미지:\n" + "\n".join(d['main_imgs']) + "\n동영상:\n\nSKU 속성 이미지:\n"
    fake_raw_text += "\n".join(d.get('sku_imgs', [])) + "\n설명 이미지:\n" + "\n".join(d['detail_imgs']) + "\n"
    for sku in d['sku_list']:
        fake_raw_text += f'"{sku["이미지"]}"\t"{sku["옵션명"]}"\t""\t"{sku["위안화"]}"\n'

    ai, error_log = analyze_with_ai(gemini_key, d['product_title'], fake_raw_text)
    if error_log: raise Exception(f"AI 분석 실패: {error_log}")
    if ai: ai['tags'] = ai.get('tags', '').replace('#', '').strip()
    else: ai = {"title": "상품명오류", "tags": "", "material": d['attributes'].get('소재', ''), "alt_text": "", "season": ""}
    
    df = pd.DataFrame(d['sku_list'])
    if df.empty: raise Exception("SKU(옵션) 데이터가 없습니다.")
    df['원가(원)'] = (df['위안화'] * ex_rate).astype(int)
    df[['최종 납품가(원)', '내 마진(원)', '쿠팡판매가(원)', '쿠팡마진율(%)', '권장소비자가(원)']] = df.apply(lambda row: optimize_profit_margin(row, target_profit), axis=1)
    calculated_skus = df.to_dict('records')
    
    for sku in calculated_skus:
        sku['옵션명'] = clean_option_name(sku['옵션명'], gemini_key, sku.get('이미지', ''))
    
    suggested_name = ai.get('title', '')
    final_material, final_tags, alt_text, season = ai.get('material', ''), ai.get('tags', ''), ai.get('alt_text', ''), ai.get('season', '')
    pack_weight, pack_size = d['attributes'].get('무게', '500'), "300*400*50"
    
    # [수술 1] "고객센터: 고객센터" 중복 제거를 위해 순수 전화번호만 입력
    importer, phone = "경기도 시흥시 비둘기공원7길 65 유한프라자 502호 삼삼오오", "070-4128-6398"
    
    output_path = os.path.join(default_out, f"[{brand_name}] {get_clean_filename(brand_name, suggested_name, '')}_최종결과물")
    os.makedirs(output_path, exist_ok=True)

    with open(os.path.join(output_path, "_raw_data_source.txt"), "w", encoding="utf-8") as f:
        f.write(fake_raw_text)
    
    columns, reqs, target_sheet_name = extract_excel_columns(target_excel_path)
    wb = openpyxl.load_workbook(target_excel_path)
    ws = wb[target_sheet_name]
    
    col_info_list = []
    for cell in ws[5]:
        if cell.value:
            c_name = str(cell.value).strip()
            c_req = str(ws.cell(row=6, column=cell.column).value).strip()
            col_info_list.append((cell.column, c_name, c_req))
            
    start_row = 9 
    needs_size_chart = any("사이즈표" in str(c) or "사이즈차트" in str(c) for c in columns)
    st.session_state.needs_size_chart = needs_size_chart

    for i, sku in enumerate(calculated_skus):
        current_row = start_row + i
        base_fname = get_clean_filename(brand_name, suggested_name, sku['옵션명'])
        
        for col_idx, col_name, col_req in col_info_list:
            val = get_excel_mapped_value(col_name, col_req, d, sku, base_fname, brand_name, suggested_name, final_tags, alt_text, season, pack_weight, pack_size, custom_category)
            if val is not None and val != "nan" and val != "":
                ws.cell(row=current_row, column=col_idx, value=val)
                
    wb.save(os.path.join(output_path, f"{os.path.splitext(os.path.basename(target_excel_path))[0]}_제출.xlsx"))
    
    base_size = create_smart_size_chart(calculated_skus) if needs_size_chart else None
    for idx, sku in enumerate(calculated_skus):
        opt_name = sku['옵션명']
        base_fname = get_clean_filename(brand_name, suggested_name, opt_name)
        c_color, c_size = split_color_size(opt_name)
        if sku['이미지']:
            di = download_img(sku['이미지'])
            if di: 
                p_main = create_studio_main_image(di) if "옵션 2" in main_strategy else trim_and_crop_1_to_1(di)
                p_main.save(os.path.join(output_path, f"{base_fname}.jpg"), format="JPEG", quality=100)
        
        display_opt = opt_name.replace("혼합색상", "").replace("단일색상", "").replace("혼합", "").replace(",", "").strip()
        display_opt = re.sub(r'\s+', ' ', display_opt)
        
        full_item_name = f"{brand_name} {suggested_name}"
        urls = get_final_detail_urls(None, d, sku, detail_strategy)
        
        p_detail = create_dynamic_detail_page(full_item_name, display_opt, final_material, c_size, "중국", urls, brand_name)
        if p_detail: p_detail.save(os.path.join(output_path, f"{base_fname}detail.jpg"), format="JPEG", quality=85)
        
        p_label = create_perfect_korean_label_900x1200(f"{full_item_name} {display_opt}".strip(), c_size, final_material, importer, phone)
        p_label.save(os.path.join(output_path, f"{base_fname}label.jpg"), format="JPEG", quality=100)
        if needs_size_chart and base_size: base_size.save(os.path.join(output_path, f"{base_fname}size.jpg"), format="JPEG", quality=95)
            
    return {"folder_path": default_out, "folder_name": os.path.basename(output_path), "excel_path": target_excel_path, "ai_result": ai, "data": d, "raw_text": fake_raw_text, "calculated_skus": calculated_skus, "output_path": output_path, "original_title": suggested_name}

# ==========================================
# 메인 UI
# ==========================================
st.title("🔬 [마스터 v10.9] 무결점 3구역 무인 공장")

tab_local, tab_api, tab_keyword = st.tabs(["📁 [무료] 로컬 폴더 모드", "🔗 [유료] API URL 모드", "🔎 [유료] AI 키워드 모드 (Phase 4)"])

with tab_local:
    st.markdown("### 🚀 [모드 A] 전체 폴더 무인 자동화")
    raw_path = st.text_input("📁 로컬 작업용 폴더 경로", value=default_local)

    col_btn1, col_btn_reset = st.columns([3, 1])
    with col_btn1:
        if st.button("🚀 전체 로컬 폴더 무인 자동화 시작", type="primary", use_container_width=True):
            st.session_state.batch_logs = []
            root_folder = fix_path(raw_path)
            found_folders = get_all_valid_folders(root_folder)
            if not found_folders: st.error("❌ 작업할 폴더를 찾을 수 없습니다.")
            else:
                st.session_state.folder_list = found_folders
                progress_bar = st.progress(0)
                status_text = st.empty()
                log_area = st.empty()
                last_success_state = None
                
                for idx, f_path in enumerate(found_folders):
                    f_name = os.path.basename(f_path)
                    status_text.info(f"🔄 [{idx+1}/{len(found_folders)}] 작업 중: {f_name} ...")
                    try:
                        state_dict = process_folder_fully_headless(f_path, root_folder, global_ex_rate, global_target_profit, global_main_strategy, global_detail_strategy)
                        st.session_state.batch_logs.append(f"✅ [{idx+1}/{len(found_folders)}] {f_name} : 성공")
                        last_success_state = state_dict
                    except Exception as e:
                        st.session_state.batch_logs.append(f"❌ [{idx+1}/{len(found_folders)}] {f_name} : 에러 패스 ({str(e)})")
                    log_area.text_area("📋 실시간 작업 로그", "\n".join(st.session_state.batch_logs), height=300)
                    progress_bar.progress((idx + 1) / len(found_folders))
                status_text.success(f"🎉 총 {len(found_folders)}개 폴더 생성 완료! (제출용 폴더를 확인하세요)")
                st.balloons()
                if last_success_state:
                    st.session_state.parsed_data = last_success_state
                    st.session_state.calculated_skus = last_success_state["calculated_skus"]

    with col_btn_reset:
        if st.button("🛑 화면/메모리 초기화", use_container_width=True):
            reset_state()
            st.rerun()

    st.markdown("---")
    st.markdown("### 🛠️ [모드 B] 최종 결과물 수동 수정 (스마트 역추적)")
    manual_folder_path = st.text_input("📂 수정할 '최종결과물 폴더' 경로 붙여넣기")
    
    if st.button("📥 작업실 열기", use_container_width=True):
        if manual_folder_path:
            target_folder = fix_path(manual_folder_path)
            def find_src_txt(p):
                for f in os.listdir(p):
                    if f.endswith(".txt") and ("_URL" in f or "_raw_data" in f): return os.path.join(p, f)
                return None
            info_file = find_src_txt(target_folder) or find_src_txt(os.path.dirname(target_folder))
            
            if not info_file: st.error("❌ 원본 데이터(.txt)를 찾을 수 없습니다.")
            else:
                try: raw_text = open(info_file, 'r', encoding='utf-8').read()
                except: raw_text = open(info_file, 'r', encoding='cp949').read()
                
                d = parse_1688_text(raw_text, MY_API_KEY)
                ai, _ = analyze_with_ai(MY_API_KEY, d['product_title'], raw_text)
                
                # 모드 B에서 이미 구워진 엑셀을 가장 먼저 찾도록 우선순위 재조정
                excel_file = None
                if os.path.exists(target_folder):
                    for f in os.listdir(target_folder):
                        if f.endswith("_제출.xlsx") and not f.startswith("~$"):
                            excel_file = os.path.join(target_folder, f)
                            break
                if not excel_file:
                    excel_file = find_excel_template(target_folder) or find_excel_template(os.path.dirname(info_file)) or find_excel_template(default_local)
                
                columns, reqs, _ = extract_excel_columns(excel_file) if excel_file else ([], [], None)
                st.session_state.needs_size_chart = any("사이즈표" in str(c) or "사이즈차트" in str(c) for c in columns)
                
                st.session_state.parsed_data = {
                    "folder_path": os.path.dirname(info_file), "excel_path": excel_file, 
                    "ai_result": ai, "data": d, "raw_text": raw_text, "output_path": target_folder,
                    "original_title": ai.get('title', '')
                }
                df = pd.DataFrame(d['sku_list'])
                df['원가(원)'] = (df['위안화'] * global_ex_rate).astype(int)
                df[['최종 납품가(원)', '내 마진(원)', '쿠팡판매가(원)', '쿠팡마진율(%)', '권장소비자가(원)']] = df.apply(lambda r: optimize_profit_margin(r, global_target_profit), axis=1)
                st.session_state.calculated_skus = df.to_dict('records')
                st.success("✅ 역추적 로드 완료! 화면 하단 탭을 확인하세요.")

with tab_api:
    st.info("💡 **API 모드:** URL 리스트 파일(`url리스트.xlsx`)과 견적서 양식들을 아래 경로에 모아두면 한 번에 추출합니다.")
    api_root_path = st.text_input("📁 API 작업용 폴더 경로", value=default_api, key="api_root_path")
    
    if st.button("🚀 URL 스캔 및 API 다이렉트 자동화 시작", type="primary", use_container_width=True):
        if not rapid_api_key: st.error("❌ 좌측 사이드바에 RapidAPI Key를 입력해주세요!")
        else:
            folder_path = fix_path(api_root_path)
            items = read_urls_from_file(folder_path)
            
            if not items: st.warning("❌ 폴더 내에서 URL 리스트 파일(.xlsx 또는 .txt)을 찾지 못했거나 URL을 찾을 수 없습니다.")
            else:
                st.session_state.batch_logs = []
                st.session_state.url_list = items
                
                progress_bar = st.progress(0)
                status_text = st.empty()
                log_area = st.empty()
                last_success_state = None
                
                for idx, item in enumerate(items):
                    target_url = item['url']
                    target_category = item['category']
                    target_template = item['template'] 
                    
                    status_text.info(f"🔄 [{idx+1}/{len(items)}] API 작업 중: {target_url} ...")
                    try:
                        state_dict = process_url_fully_headless(target_url, folder_path, target_template, rapid_api_key, MY_API_KEY, global_ex_rate, global_target_profit, global_main_strategy, global_detail_strategy, custom_category=target_category)
                        st.session_state.batch_logs.append(f"✅ [{idx+1}/{len(items)}] 성공! ({state_dict['folder_name']})")
                        last_success_state = state_dict
                    except Exception as e:
                        st.session_state.batch_logs.append(f"❌ [{idx+1}/{len(items)}] 에러: {str(e)}")
                        
                    log_area.text_area("📋 API 실시간 작업 로그", "\n".join(st.session_state.batch_logs), height=300)
                    progress_bar.progress((idx + 1) / len(items))
                
                status_text.success(f"🎉 총 {len(items)}개 상품 결과물 생성 완료! (제출용 폴더 확인)")
                st.balloons()
                
                if last_success_state:
                    st.session_state.parsed_data = last_success_state
                    st.session_state.calculated_skus = last_success_state["calculated_skus"]

with tab_keyword:
    st.markdown("## 🔍 1688 AI 키워드 스캐너 (Phase 4)")
    st.info("💡 **N사 데이터 마스터 완벽 연동 모듈.** (현재 개발 준비 중)")


# ==========================================
# 💎 가공 결과 상세 UI 
# ==========================================
if st.session_state.parsed_data:
    res = st.session_state.parsed_data
    d = res['data']
    ai = res['ai_result']
    first_opt_name = st.session_state.calculated_skus[0]['옵션명'] if st.session_state.calculated_skus else "옵션없음"
    
    output_path = res.get('output_path') if res.get('output_path') else os.path.join(default_out, f"[{brand_name}] {get_clean_filename(brand_name, ai.get('title', '상품명'), '')}_최종결과물")
    st.markdown("---")
    st.info(f"📂 결과물 저장 폴더: `{output_path}`")
    
    tab1, tab2, tab3, tab4, tab5 = st.tabs(["🤖 1. AI 기획", "💰 2. 마진 계산", "📑 3. 엑셀 굽기", "🖼️ 4. 가공 결과", "🛠️ 5. 이미지 교체소"])
    
    with tab1:
        c1, c2, c3 = st.columns([1, 1, 1])
        with c1: 
            suggested_name = st.text_input("✏️ 제안명 (수정 시 실시간 반영)", value=ai['title'])
            res['ai_result']['title'] = suggested_name 
        with c2: final_tags = st.text_input("🏷️ 검색 태그", value=ai['tags'])
        with c3: alt_text = st.text_input("📝 이미지 대체 텍스트", value=ai['alt_text'])
        
        c_attr, c_opt = st.columns([1, 2])
        with c_attr:
            final_material = st.text_input("🧵 번역된 소재", value=ai['material'])
            season = st.text_input("⛅ 계절", value=ai['season'])
            pack_size = st.text_input("📦 단품 포장 사이즈", value="300*400*50") 
            pack_weight = st.text_input("⚖️ 단품 무게 (g)", value=d['attributes'].get('무게'))
        with c_opt:
            if st.session_state.calculated_skus:
                unique_options = list(set([item['옵션명'] for item in st.session_state.calculated_skus]))
                st.text_area(f"총 {len(unique_options)}개의 깔끔한 옵션 (한자 제거됨)", value="\n".join([f"• {opt}" for opt in unique_options]), height=200, disabled=True)

    with tab2:
        cc1, cc2 = st.columns(2)
        local_ex_rate = cc1.number_input("현재 폴더 환율 (원/¥)", value=global_ex_rate, step=5)
        local_target_profit = cc2.number_input("현재 폴더 목표 순이익 (원)", value=global_target_profit, step=500)
        
        if st.session_state.calculated_skus:
            df = pd.DataFrame(st.session_state.calculated_skus)
            
            # [수술 4] 동적 재계산 로직: 데이터프레임이 렌더링될 때마다 위안화를 기반으로 전체 수식을 다시 돌립니다.
            df['원가(원)'] = (df['위안화'] * local_ex_rate).astype(int)
            df[['최종 납품가(원)', '내 마진(원)', '쿠팡판매가(원)', '쿠팡마진율(%)', '권장소비자가(원)']] = df.apply(lambda row: optimize_profit_margin(row, local_target_profit), axis=1)
            
            display_df = df.copy()
            clean_opts = display_df['옵션명'].str.replace("혼합색상", "").str.replace("단일색상", "").str.replace("혼합", "").replace(",", "").str.strip().apply(lambda x: re.sub(r'\s+', ' ', x))
            display_df.insert(0, '최종상품명', f"{brand_name} {suggested_name} " + clean_opts)
            
            edited_df = st.data_editor(
                display_df[['옵션명', '최종상품명', '위안화', '원가(원)', '최종 납품가(원)', '내 마진(원)', '쿠팡판매가(원)', '쿠팡마진율(%)', '권장소비자가(원)']], 
                use_container_width=True,
                column_config={
                    "최종상품명": st.column_config.Column(disabled=True),
                    "원가(원)": st.column_config.Column(disabled=True),
                    "최종 납품가(원)": st.column_config.Column(disabled=True),
                    "내 마진(원)": st.column_config.Column(disabled=True),
                    "쿠팡판매가(원)": st.column_config.Column(disabled=True),
                    "쿠팡마진율(%)": st.column_config.Column(disabled=True),
                    "권장소비자가(원)": st.column_config.Column(disabled=True)
                },
                key="dynamic_margin_editor"
            )
            
            if not df['위안화'].equals(edited_df['위안화']) or not df['옵션명'].equals(edited_df['옵션명']):
                df['위안화'] = edited_df['위안화']
                df['옵션명'] = edited_df['옵션명']
                st.session_state.calculated_skus = df.to_dict('records')
                st.rerun()

    with tab3:
        if res['excel_path']:
            # [수술 3] 엑셀 굽기 전 완벽한 매핑 데이터 투시경 (직접 편집 가능 Data Editor) 장착
            st.markdown("### 🔍 엑셀 매핑 데이터 표 (직접 수정 가능)")
            st.info("💡 아래 표에서 셀을 직접 더블클릭하여 데이터를 수정할 수 있습니다. [엑셀 파일 굽기] 버튼을 누르면 이 표의 내용이 그대로 엑셀에 반영됩니다.")
            
            try:
                columns, reqs, target_sheet = extract_excel_columns(res['excel_path'])
                wb = openpyxl.load_workbook(res['excel_path'])
                ws = wb[target_sheet]
                
                col_info_list = []
                for cell in ws[5]:
                    if cell.value: col_info_list.append((cell.column, str(cell.value).strip(), str(ws.cell(row=6, column=cell.column).value).strip()))
                
                preview_data = []
                target_excel_file = os.path.join(output_path, f"{os.path.splitext(os.path.basename(res['excel_path'].replace('_제출', '')))[0]}_제출.xlsx")
                
                # 기본 매핑 생성
                for sku in st.session_state.calculated_skus:
                    base_fname = get_clean_filename(brand_name, suggested_name, sku['옵션명'])
                    row_data = {"옵션명(참고용)": sku['옵션명']}
                    for col_idx, col_name, col_req in col_info_list:
                        val = get_excel_mapped_value(col_name, col_req, d, sku, base_fname, brand_name, suggested_name, final_tags, alt_text, season, pack_weight, pack_size)
                        row_data[col_name] = val if val is not None else ""
                    preview_data.append(row_data)
                
                df_preview = pd.DataFrame(preview_data)
                
                # 모드 B(스마트 역추적) 진입 시, 이미 구워져 있는 엑셀이 있다면 데이터를 우선 덮어쓰기 합니다.
                if os.path.exists(target_excel_file):
                    try:
                        ex_wb = openpyxl.load_workbook(target_excel_file, data_only=True)
                        ex_ws = ex_wb[target_sheet]
                        for i in range(len(st.session_state.calculated_skus)):
                            row_idx = 9 + i
                            for col_idx, col_name, _ in col_info_list:
                                c_val = ex_ws.cell(row=row_idx, column=col_idx).value
                                if c_val is not None:
                                    df_preview.at[i, col_name] = str(c_val)
                        st.success("✅ 결과물 폴더에 기존 생성된 엑셀 파일이 있어 성공적으로 불러왔습니다.")
                    except Exception as e:
                        st.warning(f"기존 엑셀 파일 로드 중 일부 오류가 발생했습니다: {e}")
                
                edited_preview_df = st.data_editor(df_preview, use_container_width=True, key="excel_editor")
                
                if st.button("🔥 엑셀 파일 굽기 (표 내용 반영)", type="primary"):
                    with st.spinner("엑셀 굽는 중..."):
                        try:
                            # 템플릿을 다시 불러와서, 화면의 에디터(edited_preview_df) 데이터를 물리적으로 밀어 넣습니다.
                            out_wb = openpyxl.load_workbook(res['excel_path'])
                            out_ws = out_wb[target_sheet]
                            for i, row in edited_preview_df.iterrows():
                                current_row = 9 + i
                                for col_idx, col_name, _ in col_info_list:
                                    val = row.get(col_name, "")
                                    if pd.isna(val): val = ""
                                    out_ws.cell(row=current_row, column=col_idx, value=val)
                                        
                            os.makedirs(output_path, exist_ok=True) 
                            out_wb.save(target_excel_file)
                            st.success(f"🎉 엑셀 굽기 완료! ({target_excel_file})")
                        except Exception as e: st.error(f"엑셀 에러: {e}")
            except Exception as e:
                st.error(f"엑셀 렌더링 에러: {e}")

            st.markdown("---")
            st.markdown("### 🔄 원클릭 스마트 이미지 파일명 동기화")
            original_name = res.get('original_title', '')
            
            if original_name and original_name != suggested_name:
                st.warning(f"⚠️ 상품명이 변경되었습니다: '{original_name}' -> '{suggested_name}'\n엑셀을 굽기 전 폴더 내 이미지 파일명을 동기화하세요!")
                if st.button("🚀 1초 파일명 원클릭 동기화", type="primary", use_container_width=True):
                    old_base = get_clean_filename(brand_name, original_name, "")
                    new_base = get_clean_filename(brand_name, suggested_name, "")
                    count = 0
                    try:
                        for fname in os.listdir(output_path):
                            if old_base in fname and fname.endswith(('.jpg', '.png', '.jpeg')):
                                new_fname = fname.replace(old_base, new_base)
                                os.rename(os.path.join(output_path, fname), os.path.join(output_path, new_fname))
                                count += 1
                        res['original_title'] = suggested_name 
                        st.success(f"✅ 완벽합니다! 총 {count}개의 이미지 파일명이 '{suggested_name}'(으)로 동기화되었습니다. 이제 엑셀을 다시 구워주세요!")
                    except Exception as e:
                        st.error(f"동기화 에러: {e}")
            else:
                st.info("✅ 상품명과 파일명이 완벽히 동기화된 상태입니다.")

    with tab4:
        c_opt1, c_opt2 = st.columns(2)
        with c_opt1:
            main_img_strategy = st.radio("1️⃣ 대표 이미지 가공 방식 (필수 선택)", 
                ["[옵션 1] 원본 보존 1:1 패딩", "[옵션 2] AI 배경제거 & 스튜디오"], index=0, label_visibility="collapsed")
            detail_img_strategy = st.radio("2️⃣ 상세 이미지 제작 방식", 
                ["[옵션 1] 해당 옵션 썸네일 1장 + 메인 이미지 병합", "[옵션 2] 설명 이미지 원본 무제한 쫙 이어붙이기"], index=1, label_visibility="collapsed")
        with c_opt2:
            if st.session_state.needs_size_chart: st.success("👕 **[사이즈표 컬럼 감지됨]** 스마트 사이즈표 이미지를 자동 생성합니다.")
            else: st.info("📦 **[사이즈표 컬럼 없음]** 엑셀 양식에 따라 사이즈표 이미지를 생성하지 않습니다.")
        
        importer, phone = "경기도 시흥시 비둘기공원7길 65 유한프라자 502호 삼삼오오", "070-4128-6398"
        
        b1, b2 = st.columns(2)
        with b1:
            if st.button("👁️‍🗨️ 1차: 샘플 이미지 미리보기", type="primary", use_container_width=True):
                if not st.session_state.calculated_skus: st.warning("먼저 2번 마진 탭을 확인해주세요.")
                else:
                    with st.spinner("렌더링 중..."):
                        sku = st.session_state.calculated_skus[0]
                        c_color, c_size = split_color_size(sku['옵션명'])
                        dl_img = download_img(sku['이미지'])
                        
                        main_strategy = st.session_state.get('main_strategy', '옵션 1')
                        p_main = create_studio_main_image(dl_img) if "옵션 2" in main_strategy else trim_and_crop_1_to_1(dl_img) if dl_img else None
                        
                        display_opt_name = sku['옵션명'].replace("혼합색상", "").replace("단일색상", "").replace("혼합", "").replace(",", "").strip()
                        display_opt_name = re.sub(r'\s+', ' ', display_opt_name)
                        
                        full_item_name = f"{brand_name} {suggested_name}"
                        urls = get_final_detail_urls(res.get('folder_path') if res.get('folder_path') != default_out else None, d, sku, detail_img_strategy)
                        
                        p_detail = create_dynamic_detail_page(full_item_name, display_opt_name, final_material, c_size, "중국", urls, brand_name)
                        p_label = create_perfect_korean_label_900x1200(f"{full_item_name} {display_opt_name}".strip(), c_size, final_material, importer, phone)
                        p_size = create_smart_size_chart(st.session_state.calculated_skus) if st.session_state.needs_size_chart else None
                            
                        st.session_state.preview_imgs = {'main': p_main, 'detail': p_detail, 'label': p_label, 'size': p_size}
                        st.rerun()

        with b2:
            if st.button("💾 2차: 전체 이미지 수동 다시 굽기", type="secondary", use_container_width=True):
                os.makedirs(output_path, exist_ok=True)
                tot = len(st.session_state.calculated_skus)
                pb = st.progress(0)
                stxt = st.empty()
                base_size = create_smart_size_chart(st.session_state.calculated_skus) if st.session_state.needs_size_chart else None
                
                with st.spinner("이미지 다시 굽는 중..."):
                    for idx, sku in enumerate(st.session_state.calculated_skus):
                        opt_name = sku['옵션명']
                        base_fname = get_clean_filename(brand_name, suggested_name, opt_name)
                        c_color, c_size = split_color_size(opt_name)
                        
                        stxt.text(f"[{idx+1}/{tot}] '{opt_name}' 가공 중...")
                        if sku['이미지']:
                            di = download_img(sku['이미지'])
                            if di: 
                                main_strategy = st.session_state.get('main_strategy', '옵션 1')
                                p_main = create_studio_main_image(di) if "옵션 2" in main_strategy else trim_and_crop_1_to_1(di)
                                p_main.save(os.path.join(output_path, f"{base_fname}.jpg"), format="JPEG", quality=100)
                        
                        display_opt_name = opt_name.replace("혼합색상", "").replace("단일색상", "").replace("혼합", "").replace(",", "").strip()
                        display_opt_name = re.sub(r'\s+', ' ', display_opt_name)
                        
                        full_item_name = f"{brand_name} {suggested_name}"
                        urls = get_final_detail_urls(res.get('folder_path') if res.get('folder_path') != default_out else None, d, sku, detail_img_strategy)
                        
                        sku_detail = create_dynamic_detail_page(full_item_name, display_opt_name, final_material, c_size, "중국", urls, brand_name)
                        if sku_detail: sku_detail.save(os.path.join(output_path, f"{base_fname}detail.jpg"), format="JPEG", quality=85)
                        
                        create_perfect_korean_label_900x1200(f"{full_item_name} {display_opt_name}".strip(), c_size, final_material, importer, phone, os.path.join(output_path, f"{base_fname}label.jpg"))
                        if st.session_state.needs_size_chart and base_size: base_size.save(os.path.join(output_path, f"{base_fname}size.jpg"), format="JPEG", quality=95)
                        pb.progress((idx + 1) / tot)
                st.balloons()
                st.success(f"🎉 이미지 재생성 완료!")

        st.markdown("---")
        if st.session_state.preview_imgs:
            def render_safe_popup(img_obj, title):
                st.markdown(f"**{title}**") 
                if not img_obj: st.caption("생성 생략됨"); return
                with st.container(height=350, border=True): st.image(img_obj, use_container_width=True)
            c_m, c_d, c_l, c_s = st.columns(4)
            with c_m: render_safe_popup(st.session_state.preview_imgs.get('main'), "1️⃣ 대표 이미지")
            with c_d: render_safe_popup(st.session_state.preview_imgs.get('detail'), "2️⃣ 동적 상세 이미지")
            with c_l: render_safe_popup(st.session_state.preview_imgs.get('label'), "3️⃣ 고화질 라벨")
            with c_s: render_safe_popup(st.session_state.preview_imgs.get('size'), "4️⃣ 스마트 사이즈표")

    with tab5:
        st.subheader("🛠️ 5단계: 수동 이미지 스마트 교체소")
        target_replace_dir = st.text_input("📁 교체할 폴더 경로", value=output_path)
        
        if not target_replace_dir or not os.path.exists(target_replace_dir):
            st.error("폴더를 찾을 수 없습니다. 경로를 확인해 주세요.")
        else:
            with st.expander("🚀 전체 옵션 일괄 교체", expanded=True):
                col_m_all, col_d_all = st.columns(2)
                with col_m_all: new_main_all = st.file_uploader("썸네일 업로드", type=['jpg', 'jpeg', 'png'], key="m_all")
                with col_d_all: new_detail_all = st.file_uploader("상세페이지 업로드", type=['jpg', 'jpeg', 'png'], key="d_all")
                    
                if st.button("✨ 전체 SKU 일괄 덮어쓰기", key="btn_all", type="primary"):
                    if not new_main_all and not new_detail_all: st.warning("업로드된 파일이 없습니다.")
                    else:
                        success_count = 0
                        with st.spinner("이미지 교체 중..."):
                            m_img_cache, d_img_cache = None, None
                            if new_main_all: m_img_cache = Image.open(new_main_all).convert("RGB")
                            if new_detail_all: 
                                d_img_cache = Image.open(new_detail_all).convert("RGB")
                                d_img_cache.save("temp_master_detail.jpg", format="JPEG", quality=100)
                            
                            for sku in st.session_state.calculated_skus:
                                opt_name = sku['옵션명']
                                base_fname = get_clean_filename(brand_name, suggested_name, opt_name)
                                full_item_name = f"{brand_name} {suggested_name}"
                                c_color, c_size = split_color_size(opt_name)
                                
                                if m_img_cache:
                                    process_user_thumbnail(m_img_cache).save(os.path.join(target_replace_dir, f"{base_fname}.jpg"), format="JPEG", quality=100)
                                    success_count += 1
                                if d_img_cache:
                                    display_opt_name = opt_name.replace("혼합색상", "").replace("단일색상", "").replace("혼합", "").replace(",", "").strip()
                                    display_opt_name = re.sub(r'\s+', ' ', display_opt_name)
                                    sku_detail = create_dynamic_detail_page(full_item_name, display_opt_name, final_material, c_size, "중국", ["temp_master_detail.jpg"], brand_name)
                                    sku_detail.save(os.path.join(target_replace_dir, f"{base_fname}detail.jpg"), format="JPEG", quality=85)
                                    success_count += 1
                            try: os.remove("temp_master_detail.jpg")
                            except: pass
                        st.success(f"✅ 일괄 교체 완료!")

            unique_groups = list(set([sku['옵션명'] if split_color_size(sku['옵션명'])[0] in ["혼합색상", "단일색상"] else split_color_size(sku['옵션명'])[0] for sku in st.session_state.calculated_skus]))
            for group in unique_groups:
                with st.expander(f"🎨 '{group}' 덮어쓰기", expanded=False):
                    col_m, col_d = st.columns(2)
                    with col_m: new_main = st.file_uploader("썸네일", type=['jpg', 'jpeg', 'png'], key=f"m_{group}")
                    with col_d: new_detail = st.file_uploader("상세 소스", type=['jpg', 'jpeg', 'png'], key=f"d_{group}")
                        
                    if st.button(f"✨ '{group}' 교체", key=f"btn_{group}"):
                        if not new_main and not new_detail: st.warning("파일이 없습니다.")
                        else:
                            with st.spinner("교체 중..."):
                                m_img_cache = Image.open(new_main).convert("RGB") if new_main else None
                                d_img_cache = None
                                if new_detail: 
                                    d_img_cache = Image.open(new_detail).convert("RGB")
                                    d_img_cache.save("temp_color_detail.jpg", format="JPEG", quality=100)
                                
                                for sku in st.session_state.calculated_skus:
                                    opt_name = sku['옵션명']
                                    c_color, c_size = split_color_size(opt_name)
                                    if (opt_name == group) if c_color in ["혼합색상", "단일색상"] else (c_color == group):
                                        base_fname = get_clean_filename(brand_name, suggested_name, opt_name)
                                        full_item_name = f"{brand_name} {suggested_name}"
                                        if m_img_cache: process_user_thumbnail(m_img_cache).save(os.path.join(target_replace_dir, f"{base_fname}.jpg"), format="JPEG", quality=100)
                                        if d_img_cache:
                                            display_opt_name = opt_name.replace("혼합색상", "").replace("단일색상", "").replace("혼합", "").replace(",", "").strip()
                                            display_opt_name = re.sub(r'\s+', ' ', display_opt_name)
                                            sku_detail = create_dynamic_detail_page(full_item_name, display_opt_name, final_material, c_size, "중국", ["temp_color_detail.jpg"], brand_name)
                                            sku_detail.save(os.path.join(target_replace_dir, f"{base_fname}detail.jpg"), format="JPEG", quality=85)
                                try: os.remove("temp_color_detail.jpg")
                                except: pass
                            st.success(f"✅ '{group}' 교체 완료!")